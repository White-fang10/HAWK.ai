"""
Hawk.ai Attendance Backend — main.py
=====================================
Migration: Burst-capture architecture (SCRFD + GhostFaceNet ONNX)
  - Removed: RTSP streaming, MJPEG, WebSocket, upload_frame, TemporalVoter
  - Added: /api/attendance/burst (5-frame burst POST)
  - Kept: All auth, student, analytics, session, admin, report endpoints
"""

from fastapi import (
    FastAPI, Depends, HTTPException, UploadFile, File, status
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from fastapi.security import OAuth2PasswordBearer
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import engine, Base, get_db
import models, schemas
from typing import List, Optional
import uvicorn
import asyncio
import json
import cv2
import numpy as np
from contextlib import asynccontextmanager
import datetime
import io
from pydantic import BaseModel
import os
import secrets
import bcrypt

# JWT
from jose import JWTError, jwt

# ─────────────────────────────────────────────
# Config from environment
# ─────────────────────────────────────────────
from dotenv import load_dotenv
load_dotenv()

SECRET_KEY = os.getenv("SECRET_KEY", "CHANGE_ME_IN_PRODUCTION_" + secrets.token_hex(16))
ALGORITHM  = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 8

_origins_raw = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000")
ALLOWED_ORIGINS = [o.strip() for o in _origins_raw.split(",") if o.strip()]

# ✅ Create DB tables BEFORE importing pipeline
Base.metadata.create_all(bind=engine)

from vision.pipeline import pipeline

# ─────────────────────────────────────────────
# JWT helpers
# ─────────────────────────────────────────────

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)


def create_token(user_id: int, email: str, role: str) -> str:
    """Create a signed JWT with 8h expiry."""
    expire = datetime.datetime.utcnow() + datetime.timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    payload = {"sub": str(user_id), "email": email, "role": role, "exp": expire}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    """Decode JWT and return user dict. Raises 401 on failure."""
    if not token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated",
            headers={"WWW-Authenticate": "Bearer"},
        )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return {
            "id": int(payload["sub"]),
            "email": payload["email"],
            "role": payload["role"],
        }
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid or expired token",
            headers={"WWW-Authenticate": "Bearer"},
        )


def require_admin(user: dict = Depends(get_current_user)) -> dict:
    """Require admin or super_admin role."""
    if user["role"] not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


# ─────────────────────────────────────────────
# Password helpers (bcrypt)
# ─────────────────────────────────────────────

def _hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def _verify_password(password: str, hashed: str) -> bool:
    try:
        if hashed.startswith("$2") or hashed.startswith(b"$2"):
            return bcrypt.checkpw(password.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        pass

    # Fallback for old SHA256 hashes generated prior to bcrypt migration
    import hashlib
    old_hash = hashlib.sha256(password.encode("utf-8")).hexdigest()
    return old_hash == hashed


def _seed_admin_users(db: Session):
    """Create default admin users, or upgrade SHA256 hashes to bcrypt on startup."""
    defaults = [
        ("Super Admin", "admin@hawkai.edu", "admin123", models.RoleEnum.super_admin),
        ("Dr. Sarah Mitchell", "user@hawkai.edu", "user123", models.RoleEnum.teacher),
    ]

    for name, email, password, role in defaults:
        user = db.query(models.AdminUser).filter(
            func.lower(models.AdminUser.email) == email.lower()
        ).first()
        if user is None:
            # Create fresh with bcrypt
            db.add(models.AdminUser(
                name=name, email=email,
                password_hash=_hash_password(password),
                role=role,
            ))
        elif not user.password_hash.startswith("$2b$") and not user.password_hash.startswith("$2a$"):
            # Upgrade from SHA256 -> bcrypt
            user.password_hash = _hash_password(password)

    db.commit()


# ─────────────────────────────────────────────
# Lifespan — FIXED: removed _get_scrfd import,
# using pipeline.warmup() instead
# ─────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    db = next(get_db())
    try:
        _seed_admin_users(db)
        pipeline.reload_index(db)
        # Warm up both detector and recognizer models
        # This replaces the old: from vision.detector import _get_scrfd
        pipeline.warmup()
        print("[Startup] All models ready.")
    finally:
        db.close()
    yield


app = FastAPI(title="Hawk.ai Attendance Backend", lifespan=lifespan)

# CORS — NEVER allow_origins=["*"] with credentials=True
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def read_root():
    return {"message": "Hawk.ai Backend API is running"}


# ─────────────────────────────────────────────
# HEALTH (public — no auth)
# ─────────────────────────────────────────────

@app.get("/api/health")
def health_check(db: Session = Depends(get_db)):
    total = db.query(models.Student).count()
    enrolled = db.query(models.Student).filter(models.Student.embedding.isnot(None)).count()
    with pipeline._lock:
        index_size = len(pipeline._index_ids)
    return {
        "status": "ok",
        "students_total": total,
        "students_enrolled": enrolled,
        "index_size": index_size,
        "detector": "SCRFD-2.5G ONNX",
        "recognizer": "GhostFaceNet-W1.3 ONNX",
        "mode": "burst",
    }


# Camera health (still needed by existing frontend checks)
@app.get("/api/camera/health")
def camera_health():
    return {"status": "ok", "message": "Hawk.ai backend is reachable"}


# ─────────────────────────────────────────────
# STUDENT ENDPOINTS
# ─────────────────────────────────────────────

@app.get("/api/students", response_model=List[schemas.StudentOut])
def get_students(
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    students = db.query(models.Student).order_by(models.Student.roll).all()
    return [
        schemas.StudentOut(
            id=s.id, name=s.name, roll=s.roll, email=s.email, phone=s.phone,
            avatar=s.avatar or s.name[:2].upper(),
            attendance=s.attendance_percentage,
            status=s.current_status.value if s.current_status else "absent"
        ) for s in students
    ]


@app.post("/api/students", response_model=schemas.StudentOut)
def create_student(
    student: schemas.StudentCreate,
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    from sqlalchemy.exc import IntegrityError

    if db.query(models.Student).filter(models.Student.roll == student.roll).first():
        raise HTTPException(status_code=400, detail=f"A student with roll number '{student.roll}' already exists.")
    if db.query(models.Student).filter(models.Student.email == student.email).first():
        raise HTTPException(status_code=400, detail=f"A student with email '{student.email}' is already registered.")

    db_student = models.Student(
        name=student.name,
        roll=student.roll,
        email=student.email,
        phone=student.phone or "",
        avatar=student.avatar or student.name[:2].upper(),
        attendance_percentage=0.0,
        current_status=models.StatusEnum.absent,
    )
    db.add(db_student)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="A student with this roll number or email already exists.")
    db.refresh(db_student)

    status_val = db_student.current_status.value if hasattr(db_student.current_status, "value") else str(db_student.current_status)
    return schemas.StudentOut(
        id=db_student.id, name=db_student.name, roll=db_student.roll,
        email=db_student.email, phone=db_student.phone or "",
        avatar=db_student.avatar or db_student.name[:2].upper(),
        attendance=db_student.attendance_percentage or 0.0,
        status=status_val,
    )


@app.delete("/api/students/{student_id}")
def delete_student(
    student_id: int,
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    db.delete(student)
    db.commit()
    pipeline.reload_index(db)
    return {"message": "Student deleted successfully"}


@app.post("/api/students/{student_id}/train")
async def train_student(
    student_id: int,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    _user = Depends(get_current_user),
):
    """
    Enrol a student using 1-10 face photos via SCRFD detection + GhostFaceNet embedding.
    Accepts multiple photos via field name 'files'.
    Averages all valid embeddings into a robust centroid stored in the DB.
    """
    images = []
    for f in files:
        data = np.frombuffer(await f.read(), np.uint8)
        img  = cv2.imdecode(data, cv2.IMREAD_COLOR)
        if img is not None:
            images.append(img)

    if not images:
        raise HTTPException(400, "No valid images received")

    result = await asyncio.to_thread(
        pipeline.enroll_student, student_id, images, db
    )

    if "error" in result:
        raise HTTPException(400, result["error"])

    return {
        "message":         f"Student {student_id} enrolled.",
        "photos_accepted": result["accepted"],
        "photos_failed":   result["failed"],
    }


@app.put("/api/students/{student_id}", response_model=schemas.StudentOut)
def update_student(
    student_id: int,
    student: schemas.StudentUpdate,
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    db_student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not db_student:
        raise HTTPException(status_code=404, detail="Student not found")
    for field, value in student.model_dump(exclude_unset=True).items():
        setattr(db_student, field, value)
    db.commit()
    db.refresh(db_student)
    return schemas.StudentOut(
        id=db_student.id, name=db_student.name, roll=db_student.roll,
        email=db_student.email, phone=db_student.phone,
        avatar=db_student.avatar or db_student.name[:2].upper(),
        attendance=db_student.attendance_percentage,
        status=db_student.current_status.value if db_student.current_status else "absent"
    )


# ─────────────────────────────────────────────
# ATTENDANCE — BURST ENDPOINT (new)
# ─────────────────────────────────────────────

@app.post("/api/attendance/burst")
async def process_burst(
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    _user = Depends(get_current_user),
):
    """
    Accept 5 JPEG frames captured 400ms apart from the smartboard camera.
    Runs SCRFD detection + GhostFaceNet recognition on each frame.
    Uses set-based voting (3-of-5 frames) to confirm presence.
    Returns confirmed students with vote counts.
    """
    frames = []
    for f in files:
        data = np.frombuffer(await f.read(), np.uint8)
        img  = cv2.imdecode(data, cv2.IMREAD_COLOR)
        if img is not None:
            frames.append(img)
    if not frames:
        raise HTTPException(400, "No frames decoded")
    result = await asyncio.to_thread(pipeline.process_burst, frames, db)
    return result


# ─────────────────────────────────────────────
# SUMMARY STATS
# ─────────────────────────────────────────────

@app.get("/api/analytics/summary")
def get_summary_stats(
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    students = db.query(models.Student).all()
    total = len(students)
    if total == 0:
        return {"total": 0, "present": 0, "absent": 0, "late": 0, "rate": 0.0}
    present = sum(1 for s in students if s.current_status and s.current_status.value == "present")
    late = sum(1 for s in students if s.current_status and s.current_status.value == "late")
    absent = total - present - late
    rate = round(((present + late) / total) * 100, 1)
    return {"total": total, "present": present, "absent": absent, "late": late, "rate": rate}


# ─────────────────────────────────────────────
# ANALYTICS ENDPOINTS
# ─────────────────────────────────────────────

@app.get("/api/analytics/distribution")
def get_attendance_distribution(
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    students = db.query(models.Student).all()
    present = sum(1 for s in students if s.current_status and s.current_status.value == "present")
    late    = sum(1 for s in students if s.current_status and s.current_status.value == "late")
    absent  = sum(1 for s in students if s.current_status and s.current_status.value == "absent")
    return [
        {"name": "Present", "value": present, "fill": "#219EBC"},
        {"name": "Late",    "value": late,    "fill": "#1E3A5F"},
        {"name": "Absent",  "value": absent,  "fill": "#0D1B2A"},
    ]


@app.get("/api/analytics/daily", response_model=List[schemas.DailyAttendanceOut])
def get_daily_analytics(
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    today = datetime.date.today()
    week_start = today - datetime.timedelta(days=today.weekday())
    result = []
    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    total_students = db.query(models.Student).count()

    for i in range(7):
        day_date = week_start + datetime.timedelta(days=i)
        if i >= 5:
            has_records = db.query(models.AttendanceRecord).filter(
                func.date(models.AttendanceRecord.timestamp) == day_date.isoformat()
            ).count() > 0
            if not has_records:
                continue

        present = db.query(models.AttendanceRecord).filter(
            func.date(models.AttendanceRecord.timestamp) == day_date.isoformat(),
            models.AttendanceRecord.status == models.StatusEnum.present
        ).count()
        late = db.query(models.AttendanceRecord).filter(
            func.date(models.AttendanceRecord.timestamp) == day_date.isoformat(),
            models.AttendanceRecord.status == models.StatusEnum.late
        ).count()

        if day_date <= today:
            absent = max(0, total_students - present - late)
            result.append({"day": day_names[i], "present": present, "absent": absent, "late": late})

    if not result or all(r["present"] == 0 and r["late"] == 0 for r in result):
        return [
            {"day": d, "present": 0, "absent": total_students, "late": 0}
            for d in ["Mon", "Tue", "Wed", "Thu", "Fri"]
        ]
    return result


@app.get("/api/analytics/weekly", response_model=List[schemas.WeeklyAttendanceOut])
def get_weekly_analytics(
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    today = datetime.date.today()
    total_students = db.query(models.Student).count()
    result = []

    for i in range(4):
        week_start = today - datetime.timedelta(days=today.weekday()) - datetime.timedelta(weeks=i)
        week_end = week_start + datetime.timedelta(days=6)
        week_label = f"Week {4 - i}"
        if total_students == 0:
            result.insert(0, {"week": week_label, "rate": 0.0})
            continue
        present_records = db.query(models.AttendanceRecord.student_id).filter(
            func.date(models.AttendanceRecord.timestamp) >= week_start.isoformat(),
            func.date(models.AttendanceRecord.timestamp) <= week_end.isoformat(),
            models.AttendanceRecord.status == models.StatusEnum.present
        ).distinct().count()
        rate = round((present_records / total_students) * 100, 1)
        result.insert(0, {"week": week_label, "rate": rate})

    return result


@app.get("/api/analytics/insights", response_model=List[schemas.ClassInsightOut])
def get_class_insights(
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    students = db.query(models.Student).all()
    total = len(students)
    if total == 0:
        return [{"label": "Total Students", "value": "0", "trend": "neutral"}]
    avg = round(sum(s.attendance_percentage for s in students) / total, 1)
    below75 = sum(1 for s in students if s.attendance_percentage < 75)
    top = max(students, key=lambda s: s.attendance_percentage)
    return [
        {"label": "Average Class Attendance", "value": f"{avg}%", "trend": "up"},
        {"label": "Students Below 75%", "value": str(below75), "trend": "down" if below75 > 0 else "up"},
        {"label": "Total Students", "value": str(total), "trend": "neutral"},
        {"label": "Top Attending Student", "value": top.name, "trend": "up"},
    ]


# ─────────────────────────────────────────────
# ALERTS
# ─────────────────────────────────────────────

@app.get("/api/alerts", response_model=List[schemas.AlertOut])
def get_alerts(
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    absent = db.query(models.Student).filter(
        models.Student.current_status == models.StatusEnum.absent
    ).all()
    alerts = []
    for i, s in enumerate(absent[:5], 1):
        alerts.append({"id": i, "type": "warning",
                        "message": f"{s.name} has not been detected in class",
                        "time": "just now"})
    if not alerts:
        alerts.append({"id": 1, "type": "info",
                        "message": "All students have been detected. Attendance is on track.",
                        "time": "just now"})
    below50 = db.query(models.Student).filter(models.Student.attendance_percentage < 50).all()
    for s in below50[:3]:
        alerts.append({"id": len(alerts) + 1, "type": "alert",
                        "message": f"{s.name}'s attendance dropped below 50%",
                        "time": "just now"})
    return alerts


# ─────────────────────────────────────────────
# ATTENDANCE RESET
# ─────────────────────────────────────────────

@app.post("/api/attendance/reset")
def reset_attendance(
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    db.query(models.Student).update({"current_status": models.StatusEnum.absent})
    db.commit()
    return {"message": "Attendance reset successfully. All students marked absent."}


# ─────────────────────────────────────────────
# SESSIONS
# ─────────────────────────────────────────────

@app.post("/api/sessions/start")
async def start_session(
    label: Optional[str] = None,
    class_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    # Close any currently active session
    db.query(models.AttendanceSession).filter(
        models.AttendanceSession.is_active == True
    ).update({"is_active": False, "ended_at": datetime.datetime.utcnow()})

    # Reset all students to absent
    db.query(models.Student).update({"current_status": models.StatusEnum.absent})

    # Create new session
    new_session = models.AttendanceSession(
        class_id=class_id,
        label=label or f"Session {datetime.datetime.now().strftime('%d %b %Y %H:%M')}",
        started_by=user["id"],
        is_active=True,
    )
    db.add(new_session)
    db.commit()
    db.refresh(new_session)

    return {
        "id": new_session.id,
        "label": new_session.label,
        "started_at": new_session.started_at.isoformat(),
        "message": "Session started. All students reset to absent.",
    }


@app.post("/api/sessions/{sid}/end")
async def end_session(
    sid: int,
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    session = db.query(models.AttendanceSession).filter(
        models.AttendanceSession.id == sid
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    session.is_active = False
    session.ended_at = datetime.datetime.utcnow()

    # Update final attendance percentages
    students = db.query(models.Student).all()
    for s in students:
        total = db.query(models.AttendanceRecord).filter(
            models.AttendanceRecord.student_id == s.id
        ).count()
        present = db.query(models.AttendanceRecord).filter(
            models.AttendanceRecord.student_id == s.id,
            models.AttendanceRecord.status == models.StatusEnum.present
        ).count()
        s.attendance_percentage = round((present / total) * 100, 1) if total > 0 else 0.0

    db.commit()
    return {"message": f"Session {sid} ended.", "ended_at": session.ended_at.isoformat()}


@app.get("/api/sessions/active")
def get_active_session(db: Session = Depends(get_db), _user: dict = Depends(get_current_user)):
    session = db.query(models.AttendanceSession).filter(
        models.AttendanceSession.is_active == True
    ).first()
    if not session:
        return {"active": False}
    return {
        "active": True,
        "id": session.id,
        "label": session.label,
        "started_at": session.started_at.isoformat(),
    }


# ─────────────────────────────────────────────
# REPORT HELPERS
# ─────────────────────────────────────────────

def _get_report_students(period: str, db: Session):
    today = datetime.date.today()
    if period == "daily":
        date_label = today.strftime("%d %b %Y")
        start = end = today
    elif period == "weekly":
        end = today
        start = today - datetime.timedelta(days=6)
        date_label = f"{start.strftime('%d %b')} - {end.strftime('%d %b %Y')}"
    elif period == "monthly":
        start = today.replace(day=1)
        end = today
        date_label = today.strftime("%B %Y")
    else:
        raise HTTPException(status_code=400, detail=f"Unknown period '{period}'. Use daily, weekly, or monthly.")

    students = db.query(models.Student).all()
    rows = []
    for s in students:
        present = db.query(models.AttendanceRecord).filter(
            models.AttendanceRecord.student_id == s.id,
            func.date(models.AttendanceRecord.timestamp) >= start.isoformat(),
            func.date(models.AttendanceRecord.timestamp) <= end.isoformat(),
            models.AttendanceRecord.status == models.StatusEnum.present
        ).count()
        status = s.current_status.value if s.current_status else "absent"
        rows.append({
            "name": s.name, "roll": s.roll, "email": s.email,
            "phone": s.phone or "-",
            "status": status.capitalize(),
            "attendance_pct": f"{s.attendance_percentage:.1f}%",
            "period_present": present,
        })
    return rows, date_label, f"{start.isoformat()} to {end.isoformat()}"


def _build_excel(rows, title: str, date_label: str) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Hawk.ai - {title}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="023047")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Period: {date_label}  |  Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    headers = ["#", "Name", "Roll No.", "Email", "Phone", "Status", "Attendance %"]
    header_row = 4
    header_fill = PatternFill("solid", fgColor="219EBC")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    status_colors = {"Present": "D1F0F6", "Absent": "FFE5CC", "Late": "FFF4CC"}
    for i, row in enumerate(rows, 1):
        r = header_row + i
        values = [i, row["name"], row["roll"], row["email"], row["phone"], row["status"], row["attendance_pct"]]
        bg = status_colors.get(row["status"], "FFFFFF")
        row_fill = PatternFill("solid", fgColor=bg) if row["status"] != "Present" else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if row_fill and col == 6:
                cell.fill = row_fill
            elif col == 6 and row["status"] == "Present":
                cell.fill = PatternFill("solid", fgColor="D1F0F6")

    col_widths = [5, 24, 14, 32, 16, 12, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    summary_row = header_row + len(rows) + 2
    ws.cell(row=summary_row, column=1, value=f"Total Students: {len(rows)}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=10)
    present_count = sum(1 for r in rows if r["status"] == "Present")
    ws.cell(row=summary_row, column=3, value=f"Present: {present_count}")
    ws.cell(row=summary_row, column=3).font = Font(bold=True, color="219EBC")
    absent_count = sum(1 for r in rows if r["status"] == "Absent")
    ws.cell(row=summary_row, column=5, value=f"Absent: {absent_count}")
    ws.cell(row=summary_row, column=5).font = Font(bold=True, color="FB8500")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_pdf(rows, title: str, date_label: str) -> bytes:
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 not installed.")

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_fill_color(2, 48, 71)
    pdf.rect(0, 0, 297, 28, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_xy(10, 8)
    pdf.cell(0, 10, f"Hawk.ai  |  {title}", ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_xy(10, 19)
    pdf.cell(0, 6, f"Period: {date_label}    Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}")

    pdf.set_text_color(0, 0, 0)
    pdf.ln(14)

    present_c = sum(1 for r in rows if r["status"] == "Present")
    absent_c  = sum(1 for r in rows if r["status"] == "Absent")
    late_c    = sum(1 for r in rows if r["status"] == "Late")
    rate_c    = round((present_c + late_c) / len(rows) * 100, 1) if rows else 0.0

    pdf.set_fill_color(242, 250, 253)
    pdf.set_font("Helvetica", "B", 10)
    for label, val in [("Total", str(len(rows))), ("Present", str(present_c)),
                        ("Absent", str(absent_c)), ("Late", str(late_c)), ("Rate", f"{rate_c}%")]:
        pdf.set_fill_color(242, 250, 253)
        pdf.cell(40, 10, f"{label}: {val}", border=1, align="C", fill=True, ln=0)
    pdf.ln(14)

    col_widths = [10, 55, 30, 75, 35, 25, 35]
    col_headers = ["#", "Name", "Roll No.", "Email", "Phone", "Status", "Attendance"]
    pdf.set_fill_color(33, 158, 188)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    for w, h in zip(col_widths, col_headers):
        pdf.cell(w, 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    status_colors_pdf = {
        "Present": (209, 240, 246), "Absent": (255, 229, 204), "Late": (255, 244, 204),
    }
    for i, row in enumerate(rows, 1):
        st = row["status"]
        bg = status_colors_pdf.get(st, (255, 255, 255))
        pdf.set_fill_color(*bg)
        pdf.set_text_color(30, 30, 30)
        values = [str(i), row["name"], row["roll"], row["email"], row["phone"], st, row["attendance_pct"]]
        for val, w in zip(values, col_widths):
            pdf.cell(w, 7, str(val)[:30], border=1, align="C" if w <= 35 else "L", fill=True)
        pdf.ln()

    return pdf.output()


# ─────────────────────────────────────────────
# REPORT DOWNLOAD ENDPOINTS
# ─────────────────────────────────────────────

@app.get("/api/reports/{period}/excel")
def download_excel_report(
    period: str,
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    rows, date_label, _ = _get_report_students(period, db)
    title_map = {"daily": "Daily Attendance Report", "weekly": "Weekly Attendance Report",
                 "monthly": "Monthly Attendance Report"}
    title = title_map.get(period, "Attendance Report")
    xlsx_bytes = _build_excel(rows, title, date_label)
    filename = f"hawk_ai_{period}_report_{datetime.date.today().isoformat()}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/{period}/pdf")
def download_pdf_report(
    period: str,
    db: Session = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    rows, date_label, _ = _get_report_students(period, db)
    title_map = {"daily": "Daily Attendance Report", "weekly": "Weekly Attendance Report",
                 "monthly": "Monthly Attendance Report"}
    title = title_map.get(period, "Attendance Report")
    pdf_bytes = _build_pdf(rows, title, date_label)
    filename = f"hawk_ai_{period}_report_{datetime.date.today().isoformat()}.pdf"
    return Response(
        content=bytes(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────
# AUTH ENDPOINT
# ─────────────────────────────────────────────

@app.post("/api/auth/login", response_model=schemas.LoginResponse)
def login(request: schemas.LoginRequest, db: Session = Depends(get_db)):
    email = request.email.strip().lower()
    user = db.query(models.AdminUser).filter(func.lower(models.AdminUser.email) == email).first()
    if not user or not _verify_password(request.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Account is disabled")

    role_map = {
        models.RoleEnum.super_admin: "super_admin",
        models.RoleEnum.admin: "admin",
        models.RoleEnum.teacher: "teacher",
    }
    role = role_map.get(user.role, "teacher")
    token = create_token(user.id, user.email, role)

    return schemas.LoginResponse(token=token, role=role, name=user.name, email=user.email)


# ─────────────────────────────────────────────
# ADMIN — STATS
# ─────────────────────────────────────────────

@app.get("/api/admin/stats", response_model=schemas.AdminStatsOut)
def get_admin_stats(
    db: Session = Depends(get_db),
    _user: dict = Depends(require_admin),
):
    students = db.query(models.Student).count()
    teachers = db.query(models.Teacher).count()
    classes = db.query(models.Class).count()
    all_students = db.query(models.Student).all()
    avg_att = round(sum(s.attendance_percentage for s in all_students) / len(all_students), 1) if all_students else 0.0
    return schemas.AdminStatsOut(
        students=students, teachers=teachers, classes=classes, attendance=avg_att,
        student_trend="+2.5%", class_trend="+4%", attendance_trend="+1.2%",
    )


# ─────────────────────────────────────────────
# ADMIN — TEACHER CRUD
# ─────────────────────────────────────────────

@app.get("/api/admin/teachers", response_model=List[schemas.TeacherOut])
def get_teachers(db: Session = Depends(get_db), _user: dict = Depends(require_admin)):
    return db.query(models.Teacher).all()


@app.post("/api/admin/teachers", response_model=schemas.TeacherOut)
def create_teacher(teacher: schemas.TeacherCreate, db: Session = Depends(get_db),
                   _user: dict = Depends(require_admin)):
    existing = db.query(models.Teacher).filter(models.Teacher.email == teacher.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Teacher with this email already exists")
    db_teacher = models.Teacher(**teacher.model_dump())
    db.add(db_teacher)
    db.commit()
    db.refresh(db_teacher)
    return db_teacher


@app.put("/api/admin/teachers/{teacher_id}", response_model=schemas.TeacherOut)
def update_teacher(teacher_id: int, teacher: schemas.TeacherUpdate,
                   db: Session = Depends(get_db), _user: dict = Depends(require_admin)):
    db_teacher = db.query(models.Teacher).filter(models.Teacher.id == teacher_id).first()
    if not db_teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")
    for field, value in teacher.model_dump(exclude_unset=True).items():
        setattr(db_teacher, field, value)
    db.commit()
    db.refresh(db_teacher)
    return db_teacher


@app.delete("/api/admin/teachers/{teacher_id}")
def delete_teacher(teacher_id: int, db: Session = Depends(get_db),
                   _user: dict = Depends(require_admin)):
    db_teacher = db.query(models.Teacher).filter(models.Teacher.id == teacher_id).first()
    if not db_teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")
    db.delete(db_teacher)
    db.commit()
    return {"message": "Teacher deleted successfully"}


# ─────────────────────────────────────────────
# ADMIN — CLASS CRUD
# ─────────────────────────────────────────────

@app.get("/api/admin/classes", response_model=List[schemas.ClassOut])
def get_classes(db: Session = Depends(get_db), _user: dict = Depends(require_admin)):
    classes = db.query(models.Class).all()
    result = []
    for c in classes:
        teacher_name = None
        if c.teacher_id:
            teacher = db.query(models.Teacher).filter(models.Teacher.id == c.teacher_id).first()
            teacher_name = teacher.name if teacher else None
        result.append(schemas.ClassOut(
            id=c.id, class_name=c.class_name, teacher_id=c.teacher_id,
            room_number=c.room_number, section=c.section,
            enrollment=c.enrollment, max_enrollment=c.max_enrollment,
            attendance_percentage=c.attendance_percentage, status=c.status,
            teacher_name=teacher_name, created_at=c.created_at
        ))
    return result


@app.post("/api/admin/classes", response_model=schemas.ClassOut)
def create_class(cls: schemas.ClassCreate, db: Session = Depends(get_db),
                 _user: dict = Depends(require_admin)):
    class_data = cls.model_dump(exclude={"teacher_password"})
    db_class = models.Class(**class_data)
    db.add(db_class)
    db.commit()
    db.refresh(db_class)

    if db_class.teacher_id and cls.teacher_password:
        teacher = db.query(models.Teacher).filter(models.Teacher.id == db_class.teacher_id).first()
        if teacher and teacher.email:
            existing_user = db.query(models.AdminUser).filter(models.AdminUser.email == teacher.email).first()
            if existing_user:
                existing_user.password_hash = _hash_password(cls.teacher_password)
                existing_user.name = teacher.name
            else:
                db.add(models.AdminUser(
                    name=teacher.name, email=teacher.email,
                    password_hash=_hash_password(cls.teacher_password),
                    role=models.RoleEnum.teacher, is_active=True,
                ))
            db.commit()

    teacher_name = None
    if db_class.teacher_id:
        teacher = db.query(models.Teacher).filter(models.Teacher.id == db_class.teacher_id).first()
        teacher_name = teacher.name if teacher else None
    return schemas.ClassOut(
        id=db_class.id, class_name=db_class.class_name, teacher_id=db_class.teacher_id,
        room_number=db_class.room_number, section=db_class.section,
        enrollment=db_class.enrollment, max_enrollment=db_class.max_enrollment,
        attendance_percentage=db_class.attendance_percentage, status=db_class.status,
        teacher_name=teacher_name, created_at=db_class.created_at
    )


@app.put("/api/admin/classes/{class_id}", response_model=schemas.ClassOut)
def update_class(class_id: int, cls: schemas.ClassUpdate, db: Session = Depends(get_db),
                 _user: dict = Depends(require_admin)):
    db_class = db.query(models.Class).filter(models.Class.id == class_id).first()
    if not db_class:
        raise HTTPException(status_code=404, detail="Class not found")
    update_data = cls.model_dump(exclude_unset=True, exclude={"teacher_password"})
    for field, value in update_data.items():
        setattr(db_class, field, value)
    db.commit()
    db.refresh(db_class)

    if cls.teacher_password and db_class.teacher_id:
        teacher = db.query(models.Teacher).filter(models.Teacher.id == db_class.teacher_id).first()
        if teacher and teacher.email:
            existing_user = db.query(models.AdminUser).filter(models.AdminUser.email == teacher.email).first()
            if existing_user:
                existing_user.password_hash = _hash_password(cls.teacher_password)
                db.commit()

    teacher_name = None
    if db_class.teacher_id:
        teacher = db.query(models.Teacher).filter(models.Teacher.id == db_class.teacher_id).first()
        teacher_name = teacher.name if teacher else None
    return schemas.ClassOut(
        id=db_class.id, class_name=db_class.class_name, teacher_id=db_class.teacher_id,
        room_number=db_class.room_number, section=db_class.section,
        enrollment=db_class.enrollment, max_enrollment=db_class.max_enrollment,
        attendance_percentage=db_class.attendance_percentage, status=db_class.status,
        teacher_name=teacher_name, created_at=db_class.created_at
    )


@app.delete("/api/admin/classes/{class_id}")
def delete_class(class_id: int, db: Session = Depends(get_db),
                 _user: dict = Depends(require_admin)):
    db_class = db.query(models.Class).filter(models.Class.id == class_id).first()
    if not db_class:
        raise HTTPException(status_code=404, detail="Class not found")
    db.delete(db_class)
    db.commit()
    return {"message": "Class deleted successfully"}


# ─────────────────────────────────────────────
# ADMIN — USER MANAGEMENT
# ─────────────────────────────────────────────

@app.get("/api/admin/users", response_model=List[schemas.AdminUserOut])
def get_admin_users(db: Session = Depends(get_db), _user: dict = Depends(require_admin)):
    return db.query(models.AdminUser).all()


@app.post("/api/admin/users", response_model=schemas.AdminUserOut)
def create_admin_user(user: schemas.AdminUserCreate, db: Session = Depends(get_db),
                      _user: dict = Depends(require_admin)):
    existing = db.query(models.AdminUser).filter(models.AdminUser.email == user.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="User with this email already exists")
    db_user = models.AdminUser(
        name=user.name, email=user.email, role=user.role,
        password_hash=_hash_password(user.password),
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user


@app.delete("/api/admin/users/{user_id}")
def delete_admin_user(user_id: int, db: Session = Depends(get_db),
                      _user: dict = Depends(require_admin)):
    db_user = db.query(models.AdminUser).filter(models.AdminUser.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete(db_user)
    db.commit()
    return {"message": "User deleted"}


@app.put("/api/admin/users/{user_id}/reset-password")
def reset_user_password(user_id: int, db: Session = Depends(get_db),
                         _user: dict = Depends(require_admin)):
    db_user = db.query(models.AdminUser).filter(models.AdminUser.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    new_password = secrets.token_urlsafe(12)
    db_user.password_hash = _hash_password(new_password)
    db.commit()
    return {"message": f"Password reset. New temporary password: {new_password}"}


# ─────────────────────────────────────────────
# ADMIN — SEARCH
# ─────────────────────────────────────────────

@app.get("/api/admin/search")
def admin_search(q: str = "", db: Session = Depends(get_db),
                 _user: dict = Depends(require_admin)):
    if not q or len(q) < 2:
        return {"students": [], "teachers": [], "classes": []}
    query = f"%{q.lower()}%"
    students = db.query(models.Student).filter(
        (models.Student.name.ilike(query)) | (models.Student.roll.ilike(query)) | (models.Student.email.ilike(query))
    ).limit(5).all()
    teachers = db.query(models.Teacher).filter(
        (models.Teacher.name.ilike(query)) | (models.Teacher.subject.ilike(query)) | (models.Teacher.email.ilike(query))
    ).limit(5).all()
    classes = db.query(models.Class).filter(models.Class.class_name.ilike(query)).limit(5).all()
    return {
        "students": [{"id": s.id, "name": s.name, "roll": s.roll, "type": "student"} for s in students],
        "teachers": [{"id": t.id, "name": t.name, "subject": t.subject, "type": "teacher"} for t in teachers],
        "classes": [{"id": c.id, "name": c.class_name, "type": "class"} for c in classes],
    }


# ─────────────────────────────────────────────
# MIGRATION STATUS
# ─────────────────────────────────────────────

@app.get("/api/migration/status")
def migration_status(db: Session = Depends(get_db)):
    students_with_embeddings = db.query(models.Student).filter(
        models.Student.embedding.isnot(None)
    ).count()
    return {
        "needs_reregistration": students_with_embeddings > 0,
        "registered_count": students_with_embeddings,
        "reason": (
            "Embeddings were generated by AdaFace and are incompatible with GhostFaceNet. "
            "All students must re-enroll with new face photos."
        ),
    }


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)